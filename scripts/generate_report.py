import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, 'data', 'autoprofit.db')
OUTPUT_PATH = os.path.join(BASE_DIR, 'output', 'autoprofit_report.xlsx')

def get_connection():
    return sqlite3.connect(DB_PATH)

def style_header(ws, headers, fill_color="1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

def generate_margin_sheet(wb, conn):
    ws = wb.create_sheet("Margin Analysis")
    headers = ["Model", "Brand", "Month", "List Price", "Dealer Price", "Factory Cost", "Gross Margin", "Margin %"]
    style_header(ws, headers)
    cursor = conn.cursor()
    cursor.execute(open(os.path.join(BASE_DIR, 'sql', 'queries', 'margin.sql')).read())
    for row in cursor.fetchall():
        ws.append(list(row))

def generate_monthly_sheet(wb, conn):
    ws = wb.create_sheet("Monthly Report")
    headers = ["Model", "Brand", "Segment", "Month", "Units", "Total Revenue", "Avg Discount", "Dealer Price", "Factory Cost", "Unit Margin", "Total Margin", "Total Incentives", "Net Margin"]
    style_header(ws, headers)
    cursor = conn.cursor()
    cursor.execute(open(os.path.join(BASE_DIR, 'sql', 'queries', 'monthly_report.sql')).read())
    for row in cursor.fetchall():
        ws.append(list(row))

def generate_alerts_sheet(wb, conn):
    ws = wb.create_sheet("Discount Alerts")
    headers = ["Model", "Brand", "Month", "Avg Discount", "Max Discount", "Excess"]
    style_header(ws, headers)
    cursor = conn.cursor()
    cursor.execute(open(os.path.join(BASE_DIR, 'sql', 'queries', 'alerts.sql')).read())
    for row in cursor.fetchall():
        ws.append(list(row))

def main():
    conn = get_connection()
    wb = Workbook()
    wb.remove(wb.active)

    generate_margin_sheet(wb, conn)
    generate_monthly_sheet(wb, conn)
    generate_alerts_sheet(wb, conn)

    wb.save(OUTPUT_PATH)
    conn.close()
    print(f'Report generated: {OUTPUT_PATH}')

if __name__ == '__main__':
    main()